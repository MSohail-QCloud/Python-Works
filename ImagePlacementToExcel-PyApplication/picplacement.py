# import xlsxwriter
import pandas as pd
# import xlsxwriter
# from xlrd import open_workbook
import openpyxl
import os.path

excel_ReviewFile='D:\Target Online Reviews phase#2\Inputfiles\Target-Online product reviews.xlsx'
excel_HelpingFile='D:\Target Online Reviews phase#2\Inputfiles\DPCI Helping File With Tab Information.xlsx'#DPCI Helping File With Tab Information.xlsx'


df = pd.read_excel (r'D:\Target Online Reviews phase#2\Inputfiles\DPCI Helping File With Tab Information.xlsx', sheet_name='final sheet')
df1=df.loc[:, ["Tab Index","Tab Name","DPCI 1","DPCI 2","DPCI 3","DPCI 4","DPCI 5","DPCI 6","DPCI 7","DPCI 8"]]
df2=df1[~df1["Tab Index"].isnull()]
book = openpyxl.load_workbook(excel_ReviewFile)
# book = open_workbook(excel_ReviewFile,on_demand=True)

for index, row in df2.iterrows():
    Sheetname = row['Tab Name']
    dpci1 = row['DPCI 1']
    dpci2 = row['DPCI 2']
    dpci3 = row['DPCI 3']
    dpci4 = row['DPCI 4']
    dpci5 = row['DPCI 5']
    dpci6 = row['DPCI 6']
    dpci7 = row['DPCI 7']
    dpci8 = row['DPCI 8']

    if (dpci1 == "" or dpci1 == 'nan'):
        continue
    else:
        # print(dpci1)
        # print(Sheetname)
        imageName = r'D:\Target Online Reviews phase#2\dpci pictures\{0}.jpg'.format(dpci1)
        file_exists = os.path.exists(imageName)
        # print(imageName)
        # sheet = book.get_sheet_by_name(Sheetname)
        sheet = book[Sheetname]
        # sheet=wb[Sheetname]
        # print(file_exists)
        cell_value = sheet.cell(5, 9).value
        # print(f'value of cell is {cell_value} and {dpci1} and {Sheetname}')
        if (file_exists == True and cell_value == None):
            # print(f'2value of cell is {cell_value}')
            sheet = book[Sheetname]
            # img = openpyxl.drawing.image.Image('1.png')
            img = openpyxl.drawing.image.Image(imageName)
            sheet["I5"] = dpci1
            img.height = 300
            img.width = 300
            img.anchor = 'J6'
            sheet.add_image(img)
            book.save(excel_ReviewFile)
    if (str(dpci2) == "" or str(dpci2) == "nan"):
        continue
    else:
        imageName = r'D:\Target Online Reviews phase#2\dpci pictures\{0}.jpg'.format(dpci2)
        file_exists = os.path.exists(imageName)
        cell_value = sheet.cell(10, 9).value
        if (file_exists == True and cell_value == None):
            print(f'value of cell is {cell_value}')
            sheet = book[Sheetname]
            img = openpyxl.drawing.image.Image(imageName)
            img.height = 300
            img.width = 300
            sheet["I10"] = dpci2
            img.anchor = 'J11'
            sheet.add_image(img)
            book.save(excel_ReviewFile)
    if (str(dpci3) == "" or str(dpci3) == "nan"):
        continue
    else:
        imageName = r'D:\Target Online Reviews phase#2\dpci pictures\{0}.jpg'.format(dpci3)
        file_exists = os.path.exists(imageName)
        cell_value = sheet.cell(15, 9).value
        if (file_exists == True and cell_value == None):
            sheet = book[Sheetname]
            img = openpyxl.drawing.image.Image(imageName)
            img.height = 300
            img.width = 300
            sheet["I15"] = dpci3
            img.anchor = 'J16'
            sheet.add_image(img)
            book.save(excel_ReviewFile)
    if (str(dpci4) == "" or str(dpci4) == "nan"):
        continue
    else:
        imageName = r'D:\Target Online Reviews phase#2\dpci pictures\{0}.jpg'.format(dpci4)
        file_exists = os.path.exists(imageName)
        cell_value = sheet.cell(20, 9).value
        if (file_exists == True and cell_value == None):
            sheet = book[Sheetname]
            img = openpyxl.drawing.image.Image(imageName)
            img.height = 300
            img.width = 300
            sheet["I20"] = dpci4
            img.anchor = 'J21'
            sheet.add_image(img)
            book.save(excel_ReviewFile)
    if (str(dpci5) == "" or str(dpci5) == "nan"):
        continue
    else:
        imageName = r'D:\Target Online Reviews phase#2\dpci pictures\{0}.jpg'.format(dpci5)
        file_exists = os.path.exists(imageName)
        cell_value = sheet.cell(25, 9).value
        if (file_exists == True and cell_value == None):
            sheet = book[Sheetname]
            img = openpyxl.drawing.image.Image(imageName)
            img.height = 300
            img.width = 300
            sheet["I25"] = dpci5
            img.anchor = 'J26'
            sheet.add_image(img)
            book.save(excel_ReviewFile)
    if (str(dpci6) == "" or str(dpci6) == "nan"):
        continue
    else:
        imageName = r'D:\Target Online Reviews phase#2\dpci pictures\{0}.jpg'.format(dpci6)
        file_exists = os.path.exists(imageName)
        cell_value = sheet.cell(30, 9).value
        if (file_exists == True and cell_value == None):
            sheet = book[Sheetname]
            img = openpyxl.drawing.image.Image(imageName)
            img.height = 300
            img.width = 300
            sheet["I30"] = dpci6
            img.anchor = 'J31'
            sheet.add_image(img)
            book.save(excel_ReviewFile)

    if (str(dpci7) == "" or str(dpci7) == "nan"):
        continue
    else:
        imageName = r'D:\Target Online Reviews phase#2\dpci pictures\{0}.jpg'.format(dpci7)
        file_exists = os.path.exists(imageName)
        cell_value = sheet.cell(35, 9).value
        if (file_exists == True and cell_value == None):
            sheet = book[Sheetname]
            img = openpyxl.drawing.image.Image(imageName)
            img.height = 300
            img.width = 300
            sheet["I35"] = dpci7
            img.anchor = 'J36'
            sheet.add_image(img)
            book.save(excel_ReviewFile)
    if (str(dpci8) == "" or str(dpci8) == "nan"):
        continue
    else:
        imageName = r'D:\Target Online Reviews phase#2\dpci pictures\{0}.jpg'.format(dpci8)
        file_exists = os.path.exists(imageName)
        cell_value = sheet.cell(40, 9).value
        if (file_exists == True and cell_value == None):
            sheet = book[Sheetname]
            img = openpyxl.drawing.image.Image(imageName)
            img.height = 100
            img.width = 100
            sheet["I40"] = dpci8
            img.anchor = 'J41'
            sheet.add_image(img)
            book.save(excel_ReviewFile)
input("Task completed.")